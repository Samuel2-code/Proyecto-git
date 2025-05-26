import os
import re
import pdfplumber
import pandas as pd
from flask import Flask, render_template, request, send_file, jsonify
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from werkzeug.utils import secure_filename
import tempfile

app = Flask(__name__)

# Expresiones regulares centralizadas
REGEX_PATTERNS = {
    'Paciente': r'Paciente\s*(.*?)\s*Identificación',
    'Edad': r'Edad\s*(\d+)\s*años',
    'Fecha': r'Atención\s*(.*?)\s*Motivo',
    'Identificación': r'Identificación\s*(.*?)\s*Teléfono',
    'Edad de Pareja': r'(?:pareja tiene|Edad:)\s*(\d+)\s*años',
    'Peso': r'xmin.\s*([\d,.]+)\s*Talla',
    'Talla': r'Talla:.\s*([\d,.]+)\s*mt',
    'Triglicerios': r'TRIGLICERIDOS:?\s*([\d,.]+)',
    'Colesterol': r'COLESTEROL TOTAL ?:?\s*([\d,.]+)',
    'HDL': r'HDL:?\s*([\d,.]+)',
    'FSH': r'FSH:?\s*([\d,.]+)',
    'LH': r'LH:?\s*([\d,.]+)',
    'Estradiol': r'ESTRADIOL:?\s*([\d,.]+)',
    'Prolactina': r'PROLACTINA:?\s*([\d,.]+)',
    'Hormona Antimulleriana': r'HORMONA ANTIMULLERIANA.*?:?\s*([\d.,-: /]+)',
    'DHEA-S': r'DEHIDROEPIANDROSTERONA SULFATO.*?:?\s*([\d,.]+)',
    'Testosterona Libre': r'TESTOSTERONA LIBRE:?\s*([\d,.]+)',
    'Testosterona Total': r'TESTOSTERONA TOTAL:?\s*([\d,.]+)',
    'Insulina Libre': r'INSULINA LIBRE:? :?\s*([\d,.]+)',
    'Glicemia': r'GLICEMIA:?\s*([\d,.]+)',
    'Hemoglobina Glicosilada': r'HEMOGLOBINA GLICOSILADA:?\s*([\d,.]+)',
    'TSH': r'TSH:?\s*([\d.,*]+)',
    'Vivos Pareja(ESPERMOGRAMA)': r'Vivos: ?\s*([\d.%*]+)',
    'Caida de Cabello': r'Caida del cabello: ?\s*(.*)',
    'Planificacion': r'PLANIFICACION: ?\s*(.*)',
    'Dismenorrea': r'Dismenorrea: ?\s*(.*)',
    'Ciclos Menstruales': r'CICLOS: ?\s*(.*)',
    'Tratamiento Fertilidad': r'Tratamientos previos de fertilidad:?\s*(.*)',
    'Infertilidad': r'Infertilidad:?\s*(.*)',
    'Acne': r'Acne:?\s*(.*)',
    'Miomatosis': r'Miomatosis ?\s*(.*)'
}

def extract_data_from_pdf(pdf_path):
    """
    Extrae datos del PDF leyendo todo el texto completo y realizando búsqueda una sola vez,
    retornando un diccionario con los datos extraídos o vacío si no se encuentra nada.
    """
    try:
        with pdfplumber.open(pdf_path) as pdf:
            full_text = ""
            # Concatenar texto de todas las páginas
            for page in pdf.pages:
                text = page.extract_text()
                if text:
                    full_text += text + "\n"  # separador opcional

            if not full_text.strip():
                return {}

            record = {}
            # Buscar cada patrón solo una vez en el texto completo
            for key, pattern in REGEX_PATTERNS.items():
                match = re.search(pattern, full_text, re.IGNORECASE)
                if match:
                    record[key] = match.group(1).strip()

            return record
    except Exception as e:
        print(f"Error procesando {pdf_path}: {e}")
        return {}

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/process', methods=['POST'])
def process():
    if 'files[]' not in request.files:
        return jsonify({'status': 'error', 'message': 'No files part in request'})

    files = request.files.getlist('files[]')
    all_records = []

    with tempfile.TemporaryDirectory() as temp_dir:
        for f in files:
            filename = secure_filename(f.filename)
            if not filename.lower().endswith('.pdf'):
                continue
            temp_path = os.path.join(temp_dir, filename)
            f.save(temp_path)
            data = extract_data_from_pdf(temp_path)
            if data:
                data['Archivo'] = filename  # Añadir nombre del archivo
                all_records.append(data)

        if not all_records:
            return jsonify({'status': 'error', 'message': 'No data extracted'})

        df = pd.DataFrame(all_records)

        # Eliminar duplicados basados en 'Paciente' y 'Archivo' para mayor certeza
        subset_cols = []
        if 'Paciente' in df.columns:
            subset_cols.append('Paciente')
        if 'Archivo' in df.columns:
            subset_cols.append('Archivo')

        if subset_cols:
            df = df.drop_duplicates(subset=subset_cols)
        else:
            df = df.drop_duplicates()

        output_path = os.path.join('static', 'datos_pacientes.xlsx')
        os.makedirs('static', exist_ok=True)
        df.to_excel(output_path, index=False)

        wb = load_workbook(output_path)
        ws = wb.active
        yellow_fill = PatternFill(start_color='a983ff', end_color='a983ff', fill_type='solid')
        for cell in ws[1]:
            cell.fill = yellow_fill
        wb.save(output_path)

    return jsonify({'status': 'success', 'file': output_path, 'rows': len(df)})

@app.route('/download')
def download():
    output_path = os.path.join('static', 'datos_pacientes.xlsx')
    if os.path.exists(output_path):
        return send_file(output_path, as_attachment=True)
    else:
        return "Archivo no encontrado", 404

if __name__ == '__main__':
    app.run(debug=True)


